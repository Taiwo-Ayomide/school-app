from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
# from openpyxl import workbook
from openpyxl import Workbook
import pathlib





class MyGUI:
    def __init__(self):

        background="#06283D"
        framebg="#EDEDED"
        framebg="#06283D"

      
        self.root = Tk()
        self.root.title("Workers Registration System")
        self.root.geometry("1250x700")
        self.root.config(bg=background)
        self.root.resizable(False, False)

        file = pathlib.Path('Student_data.xlsx')
        if file.exists():
            pass

        else:
            file = Workbook()
            sheet = file.active
            sheet['A1']="Registration  No."
            sheet['B1']="Name"
            sheet['C1']="Class"
            sheet['D1']="Gender"
            sheet['E1']="DOB"
            sheet['F1']="Date of Registration"
            sheet['G1']="Religion"
            sheet['H1']="Skill"
            sheet['I1']="Father Name"
            sheet['J1']="Mother Name"
            sheet['K1']="Father's Occupation"
            sheet['L1']="Mother's Occupation"

            file.save('Student_data.xlsx')
                #function to show Image
        
        
        #Exit Window
        def Exit():
            self.root.destroy()


        def showimage():
            global filename
            global img
            filename=filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Image file", filetype=(("JPG File","*.jpg"),("PNG File", "*.png"), ("All Files", "*.txt")))

            img = (Image.open(filename))
            resized_image = img.resize((190, 190))
            photo2 = ImageTk.PhotoImage(resized_image)
            new.config(image=photo2)
            new.image=photo2



            ###REGISTRATION NUMBER
            # We will need to design automatic registration no. entry system. I t will be created to automatic enter registration no.
        def registration_no():
            file=openpyxl.load_workbook('Student_data.xlsx')
            sheet=file.active

            row=sheet.max_row

            max_row_value=sheet.cell(row=row, column=1).value
            

            try:
                Registration.set(max_row_value+1)
            except:
                Registration.set("1")
        

        # CLEAR
        def Clear():
            global img
            Name.set('')
            DOB.set('')
            Religion.set('')
            Skill.set('')
            F_Name.set('')
            M_Name.set('')
            Father_Occupation.set('')
            Mother_Occupation.set('')
            Class.set("Select Class")

            registration_no()

            Button.config(state = 'normal')
            img=PhotoImage(file="Images/passport.png")
            new.config(image=img)
            new.image = img


            img = ""
        
        
        ##########SAVE
        def save():
            R1 = Registration.get()
            N1 = Name.get()
            C1 = Class.get()
            try:
                G1=gender
            except:
                messagebox.showerror("error", "Select Gender!")

            D2=DOB.get()
            D1=Date.get()
            Re1 = Religion.get()
            S1=Skill.get()
            fathername=F_Name.get()
            mothername=M_Name.get()
            F1 = Father_Occupation.get()
            M1 = Mother_Occupation.get()

            if N1=="" or C1 =="Select Class" or D2=="" or Re1 == "" or S1 == "" or fathername=="" or mothername=="" or F1=="" or M1=="":
                messagebox.showerror("error", "Few Data is missing!")
            else:
                file = openpyxl.load_workbook('Student_data.xlsx')
                sheet = file.active
                sheet.cell(column=1, row=sheet.max_row+1, value=R1)
                sheet.cell(column=2, row=sheet.max_row, value=N1)
                sheet.cell(column=3, row=sheet.max_row, value=C1)
                sheet.cell(column=4, row=sheet.max_row, value=G1)
                sheet.cell(column=5, row=sheet.max_row, value=D2)
                sheet.cell(column=6, row=sheet.max_row, value=D1)
                sheet.cell(column=7, row=sheet.max_row, value=Re1)
                sheet.cell(column=8, row=sheet.max_row, value=S1)
                sheet.cell(column=9, row=sheet.max_row, value=fathername)
                sheet.cell(column=10, row=sheet.max_row, value=mothername)
                sheet.cell(column=11, row=sheet.max_row, value=F1)
                sheet.cell(column=12, row=sheet.max_row, value=M1)

                file.save('Student_data.xlsx')

                try:
                    img.save("Student Images/"+str(R1)+".jpg")

                except:
                    messagebox.showinfo("info", "Profile Picture is not available!!!")
                
           
                messagebox.showinfo("info", "Successfully data entered")

                Clear()

                registration_no() # recheck reg no and reissue new

            
        #gender3
        def selection():
            global gender
            value=radio.get()
            if value== 1:
                gender = "Male"

            else:
                gender = "Female"
        

        #TOP FRAMES
        self.mail=Label(self.root, text="Email: taiwoayomide202@gmail.com", width=10, height=3, bg="#f0687c", anchor='e')
        self.mail.pack(side=TOP, fill=X)
        self.register=Label(self.root, text="WORKERS REGISTRATION", width=10, height=2, bg="#c36464",fg="#fff", font='Arial 20 bold')
        self.register.pack(side=TOP, fill=X)
        

        # Search box to update
        Search = StringVar()
        Entry(self.root, textvariable=Search, width=15, bd=2, font='Arial 20').place(x=820, y=70)
        imageicon3 = PhotoImage(file="Images/research.png")
        Srch = Button(self.root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="Arial 13 bold")
        Srch.place(x=1060, y=71)


        imageicon4=PhotoImage(file="Images/update.png")
        Update_button=Button(self.root, image=imageicon4, bg="#c36464")
        Update_button.place(x=110, y=64)


        #Registration and Date
        Label(self.root, text="Registration No:", font="Arial 13", fg="#fff", bg=background).place(x=30, y=150)
        Label(self.root, text="Date:", font="Arial 13", fg="#fff", bg=background).place(x=500, y=150)

        Registration=IntVar()
        Date = StringVar()

        reg_entry = Entry(self.root, textvariable=Registration, width=15, font="Arial 10")
        reg_entry.place(x=160, y=150)

        registration_no()

        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        date_entry = Entry(self.root, textvariable=Date, width=15,font="Arial 10")
        date_entry.place(x=550, y=150)

        Date.set(d1)


        #Student details
        obj=LabelFrame(self.root, text="Student's Detail", font=20, bd=2, width=900, bg="#fff", fg=framebg, height=250, relief=GROOVE)
        obj.place(x=30, y=200)
        
        Label(obj, text="Full Name: ", font="Arial 13", bg="#fff", fg=framebg).place(x=30, y=50)
        Label(obj, text="Date of Birth: ", font="Arial 13", bg="#fff", fg=framebg).place(x=30, y=100)
        Label(obj, text="Gender: ", font="Arial 13", bg="#fff", fg=framebg).place(x=30, y=150)

        Label(obj, text="Class: ", font="Arial 13", bg="#fff", fg=framebg).place(x=500, y=50)
        Label(obj, text="Religion: ", font="Arial 13", bg="#fff", fg=framebg).place(x=500, y=100)
        Label(obj, text="Skills: ", font="Arial 13", bg="#fff", fg=framebg).place(x=500, y=150)


        Name = StringVar()
        name_entry = Entry(obj, textvariable=Name, width=20, font="Arial 10")
        name_entry.place(x=160, y=50)

        DOB = StringVar()
        dob_entry = Entry(obj, textvariable=DOB, width=20, font="Arial 10")
        dob_entry.place(x=160, y=100)


        radio = IntVar()
        R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg="#fff",fg=framebg, command=selection)
        R1.place(x=150, y=150)

        R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg="#fff",fg=framebg, command=selection)
        R2.place(x=200, y=150)

        Religion = StringVar()
        religion_entry = Entry(obj, textvariable=Religion, width=20, font="Arial 10")
        religion_entry.place(x=630, y=100)

        Skill = StringVar()
        skill_entry = Entry(obj, textvariable=Skill, width=20, font="Arial 10")
        skill_entry.place(x=630, y=150)

        Class = Combobox(obj, values=['1', '2', '3', '4', '5','6','7','8','9','10','11','12'], font="Roboto 10", width=17, state="r")
        Class.place(x=630, y=50)
        Class.set("Select Class")

        #Parent details
        obj2=LabelFrame(self.root, text="Parent's Details", font=20, bd=2, width=900, bg="#fff", fg=framebg, height=220, relief=GROOVE)
        obj2.place(x=30, y=470)

        Label(obj2, text="Father's Name", font="Arial 13", bg="#fff", fg=framebg).place(x=30, y=50)
        Label(obj2, text="Occupation:", font="Arial 13", bg="#fff", fg=framebg).place(x=30, y=100)

        F_Name=StringVar()
        f_entry = Entry(obj2, textvariable=F_Name, width=20, font="Arial 10")
        f_entry.place(x=160, y=50)

        Father_Occupation=StringVar()
        FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="Arial 10")
        FO_entry.place(x=160, y=100)


        Label(obj2, text="Mother's Name", font="Arial 13", bg="#fff", fg=framebg).place(x=500, y=50)
        Label(obj2, text="Occupation:", font="Arial 13", bg="#fff", fg=framebg).place(x=500, y=100)

        M_Name=StringVar()
        M_entry = Entry(obj2, textvariable=M_Name, width=20, font="Arial 10")
        M_entry.place(x=630, y=50)

        Mother_Occupation=StringVar()
        MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="Arial 10")
        MO_entry.place(x=630, y=100)


        #image
        f=Frame(self.root, bd=3, bg="black", width=180, height=200, relief=GROOVE)
        f.place(x=1000, y=150)

        img=PhotoImage(file="Images/passport.png")
        new=Label(f, image=img)
        new.place(x=0, y=0)

        #button

        Button(self.root, text="Upload", width=19, height=2, font="Arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)

        saveButton=Button(self.root, text="Save", width=19, height=2, font="Arial 12 bold", bg="lightgreen", command=save).place(x=1000, y=450)

        clearbtn=Button(self.root, text="Reset", width=19, height=2, font="Arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)

        Button(self.root, text="Exit", width=19, height=2, font="Arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)



        self.root.mainloop()
MyGUI()